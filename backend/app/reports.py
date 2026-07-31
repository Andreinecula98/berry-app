"""Report generation helpers for admin exports (Excel and PDF)."""
import io
from typing import Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import Submission

STATUS_LABELS = {
    "pending": "Pending",
    "approved": "Approved",
    "rejected": "Rejected",
}

HEADERS = [
    "Field",
    "Team Leader",
    "Date",
    "Time",
    "Meter #",
    "Orange fruit",
    "White/pink fruit",
    "White fruit",
    "Big green fruit",
    "Small green fruit",
    "Opened flowers",
    "Buds",
    "Status",
    "Submitted At",
    "Reviewed By",
]


def _status_label(sub: Submission) -> str:
    status_value = sub.status.value if hasattr(sub.status, "value") else sub.status
    return STATUS_LABELS.get(status_value, status_value)



def _rows_for(sub: Submission) -> list[list]:
    reviewed_by = (sub.reviewed_by.full_name or sub.reviewed_by.username) if sub.reviewed_by else "-"
    time_value = sub.time.strftime("%H:%M") if sub.time else ""
    submitted_at = sub.created_at.strftime("%Y-%m-%d %H:%M") if sub.created_at else ""
    team_leader = sub.employee.full_name or sub.employee.username

    return [
        [
            sub.field.name,
            team_leader,
            sub.date.isoformat(),
            time_value,
            meter.meter_number,
            meter.orange_fruit,
            meter.white_pink_fruit,
            meter.white_fruit,
            meter.big_green_fruit,
            meter.small_green_fruit,
            meter.opened_flowers,
            meter.buds,
            _status_label(sub),
            submitted_at,
            reviewed_by,
        ]
        for meter in sub.meters
    ]



def build_excel_report(submissions: Sequence[Submission]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily TL Counts"

    ws.append(HEADERS)
    header_fill = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for sub in submissions:
        for row in _rows_for(sub):
            ws.append(row)

    widths = [24, 20, 12, 10, 9, 12, 15, 12, 15, 15, 15, 10, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()



def build_pdf_report(submissions: Sequence[Submission]) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=0.7 * cm,
        rightMargin=0.7 * cm,
        topMargin=0.9 * cm,
        bottomMargin=0.9 * cm,
    )
    styles = getSampleStyleSheet()
    elements = [Paragraph("Daily TL Counts Report", styles["Title"]), Spacer(1, 0.35 * cm)]

    rows = []
    for sub in submissions:
        rows.extend(_rows_for(sub))
    data = [HEADERS] + rows
    if len(data) == 1:
        data.append(["-"] * len(HEADERS))

    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6A1B9A")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 6),
                ("LEADING", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#d1d5db")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f8")]),
                ("ALIGN", (2, 0), (13, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()
