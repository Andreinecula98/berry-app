import io

from openpyxl import load_workbook


FIELD_PAYLOAD = {"name": "TL TEST FIELD", "total_meters": 120}
FIELD_UPDATE_PAYLOAD = {"name": "TL TEST FIELD UPDATED", "total_meters": 150}
METER_PAYLOAD = [
    {
        "meter_number": 1,
        "orange_fruit": 3,
        "white_pink_fruit": 4,
        "white_fruit": 5,
        "big_green_fruit": 6,
        "small_green_fruit": 7,
        "opened_flowers": 8,
        "buds": 9,
    },
    {
        "meter_number": 2,
        "orange_fruit": 10,
        "white_pink_fruit": 11,
        "white_fruit": 12,
        "big_green_fruit": 13,
        "small_green_fruit": 14,
        "opened_flowers": 15,
        "buds": 16,
    },
]

FOUR_METER_PAYLOAD = [
    {
        "meter_number": index + 1,
        "orange_fruit": index + 1,
        "white_pink_fruit": index + 2,
        "white_fruit": index + 3,
        "big_green_fruit": index + 4,
        "small_green_fruit": index + 5,
        "opened_flowers": index + 6,
        "buds": index + 7,
    }
    for index in range(4)
]



def test_root_ok(client):
    resp = client.get("/")
    assert resp.status_code == 200
    assert resp.json()["status"] == "ok"



def test_admin_login_success(admin_token):
    assert admin_token



def test_login_wrong_password(client):
    resp = client.post("/auth/login", json={"username": "admin", "password": "wrong"})
    assert resp.status_code == 401



def test_admin_can_create_employee(client, admin_token):
    resp = client.post(
        "/admin/users",
        json={"username": "ion_test", "password": "parola123", "full_name": "Ion Test", "role": "employee"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 201
    assert resp.json()["role"] == "employee"



def test_duplicate_employee_rejected(client, admin_token):
    resp = client.post(
        "/admin/users",
        json={"username": "ion_test", "password": "parola123", "role": "employee"},
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert resp.status_code == 400


def test_admin_can_update_field(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    create_resp = client.post(
        "/admin/fields",
        json={"name": "FIELD UPDATE ORIGINAL", "total_meters": 100},
        headers=admin_headers,
    )
    assert create_resp.status_code == 201
    field_id = create_resp.json()["id"]

    other_resp = client.post(
        "/admin/fields",
        json={"name": "FIELD UPDATE CONFLICT", "total_meters": 200},
        headers=admin_headers,
    )
    assert other_resp.status_code == 201

    update_resp = client.put(f"/admin/fields/{field_id}", json=FIELD_UPDATE_PAYLOAD, headers=admin_headers)
    assert update_resp.status_code == 200
    assert update_resp.json()["name"] == FIELD_UPDATE_PAYLOAD["name"]
    assert update_resp.json()["total_meters"] == FIELD_UPDATE_PAYLOAD["total_meters"]

    fields_resp = client.get("/fields", headers=admin_headers)
    assert fields_resp.status_code == 200
    updated_field = next(field for field in fields_resp.json() if field["id"] == field_id)
    assert updated_field["name"] == FIELD_UPDATE_PAYLOAD["name"]
    assert updated_field["total_meters"] == FIELD_UPDATE_PAYLOAD["total_meters"]

    conflict_resp = client.put(
        f"/admin/fields/{field_id}",
        json={"name": "FIELD UPDATE CONFLICT", "total_meters": 175},
        headers=admin_headers,
    )
    assert conflict_resp.status_code == 400
    assert conflict_resp.json()["detail"] == "This field name already exists"

    missing_resp = client.put(
        "/admin/fields/999999",
        json={"name": "MISSING FIELD", "total_meters": 99},
        headers=admin_headers,
    )
    assert missing_resp.status_code == 404
    assert missing_resp.json()["detail"] == "Field not found"


def test_admin_can_delete_unused_field(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    create_resp = client.post(
        "/admin/fields",
        json={"name": "FIELD DELETE UNUSED", "total_meters": 90},
        headers=admin_headers,
    )
    assert create_resp.status_code == 201
    field_id = create_resp.json()["id"]

    delete_resp = client.delete(f"/admin/fields/{field_id}", headers=admin_headers)
    assert delete_resp.status_code == 204

    fields_resp = client.get("/fields", headers=admin_headers)
    assert fields_resp.status_code == 200
    assert all(field["id"] != field_id for field in fields_resp.json())

    missing_resp = client.delete("/admin/fields/999999", headers=admin_headers)
    assert missing_resp.status_code == 404
    assert missing_resp.json()["detail"] == "Field not found"


def test_admin_cannot_delete_field_with_submissions(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    field_resp = client.post(
        "/admin/fields",
        json={"name": "FIELD WITH SUBMISSIONS", "total_meters": 140},
        headers=admin_headers,
    )
    assert field_resp.status_code == 201
    field_id = field_resp.json()["id"]

    employee_resp = client.post(
        "/admin/users",
        json={"username": "field_submitter", "password": "parola123", "role": "employee"},
        headers=admin_headers,
    )
    assert employee_resp.status_code == 201

    login_resp = client.post("/auth/login", json={"username": "field_submitter", "password": "parola123"})
    assert login_resp.status_code == 200
    emp_headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    submit_resp = client.post(
        "/submissions",
        json={
            "field_id": field_id,
            "date": "2026-08-02",
            "meters": [METER_PAYLOAD[0]],
        },
        headers=emp_headers,
    )
    assert submit_resp.status_code == 201

    delete_resp = client.delete(f"/admin/fields/{field_id}", headers=admin_headers)
    assert delete_resp.status_code == 400
    assert delete_resp.json()["detail"] == "Cannot delete a field that has existing submissions"


def test_field_admin_endpoints_require_admin(client, admin_token):
    create_no_auth = client.post("/admin/fields", json={"name": "FIELD NO AUTH", "total_meters": 75})
    assert create_no_auth.status_code == 401

    admin_headers = {"Authorization": f"Bearer {admin_token}"}
    field_resp = client.post(
        "/admin/fields",
        json={"name": "FIELD AUTH TARGET", "total_meters": 88},
        headers=admin_headers,
    )
    assert field_resp.status_code == 201
    field_id = field_resp.json()["id"]

    employee_resp = client.post(
        "/admin/users",
        json={"username": "field_non_admin", "password": "parola123", "role": "employee"},
        headers=admin_headers,
    )
    assert employee_resp.status_code == 201

    login_resp = client.post("/auth/login", json={"username": "field_non_admin", "password": "parola123"})
    assert login_resp.status_code == 200
    emp_headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    create_forbidden = client.post(
        "/admin/fields",
        json={"name": "FIELD FORBIDDEN CREATE", "total_meters": 95},
        headers=emp_headers,
    )
    assert create_forbidden.status_code == 403

    update_no_auth = client.put(
        f"/admin/fields/{field_id}",
        json={"name": "FIELD NO AUTH UPDATE", "total_meters": 91},
    )
    assert update_no_auth.status_code == 401

    update_forbidden = client.put(
        f"/admin/fields/{field_id}",
        json={"name": "FIELD FORBIDDEN UPDATE", "total_meters": 91},
        headers=emp_headers,
    )
    assert update_forbidden.status_code == 403

    delete_no_auth = client.delete(f"/admin/fields/{field_id}")
    assert delete_no_auth.status_code == 401

    delete_forbidden = client.delete(f"/admin/fields/{field_id}", headers=emp_headers)
    assert delete_forbidden.status_code == 403


def test_employee_login_and_submission_flow(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    field_resp = client.post("/admin/fields", json=FIELD_PAYLOAD, headers=admin_headers)
    assert field_resp.status_code == 201
    field_id = field_resp.json()["id"]

    all_fields_resp = client.get("/fields", headers=admin_headers)
    assert all_fields_resp.status_code == 200
    assert any(field["name"] == FIELD_PAYLOAD["name"] for field in all_fields_resp.json())

    login_resp = client.post("/auth/login", json={"username": "ion_test", "password": "parola123"})
    assert login_resp.status_code == 200
    emp_token = login_resp.json()["access_token"]
    emp_headers = {"Authorization": f"Bearer {emp_token}"}

    employee_fields_resp = client.get("/fields", headers=emp_headers)
    assert employee_fields_resp.status_code == 200
    assert len(employee_fields_resp.json()) >= 1

    submit_resp = client.post(
        "/submissions",
        json={
            "field_id": field_id,
            "date": "2026-07-31",
            "time": "07:15:00",
            "meters": METER_PAYLOAD,
        },
        headers=emp_headers,
    )
    assert submit_resp.status_code == 201
    body = submit_resp.json()
    assert body["field"]["name"] == FIELD_PAYLOAD["name"]
    assert body["status"] == "pending"
    assert len(body["meters"]) == 2
    assert body["meters"][1]["buds"] == 16
    assert body["averages"] == {
        "orange_fruit": 6.5,
        "white_pink_fruit": 7.5,
        "white_fruit": 8.5,
        "big_green_fruit": 9.5,
        "small_green_fruit": 10.5,
        "opened_flowers": 11.5,
        "buds": 12.5,
    }
    assert body["overall_average"] == 66.5

    mine_resp = client.get("/submissions/me", headers=emp_headers)
    assert mine_resp.status_code == 200
    assert len(mine_resp.json()) == 1
    assert mine_resp.json()[0]["meters"][0]["meter_number"] == 1
    assert mine_resp.json()[0]["averages"]["orange_fruit"] == 6.5

    forbidden = client.get("/admin/submissions", headers=emp_headers)
    assert forbidden.status_code == 403

    all_resp = client.get("/admin/submissions", headers=admin_headers)
    assert all_resp.status_code == 200
    submission = all_resp.json()[0]
    submission_id = submission["id"]
    assert submission["employee"]["username"] == "ion_test"
    assert submission["field"]["name"] == FIELD_PAYLOAD["name"]
    assert submission["averages"]["buds"] == 12.5
    assert submission["overall_average"] == 66.5

    review_resp = client.patch(
        f"/admin/submissions/{submission_id}", json={"status": "approved"}, headers=admin_headers
    )
    assert review_resp.status_code == 200
    assert review_resp.json()["status"] == "approved"
    assert review_resp.json()["reviewed_by"]["username"] == "admin"


def test_submission_rejects_more_than_three_meters(client):
    login_resp = client.post("/auth/login", json={"username": "ion_test", "password": "parola123"})
    assert login_resp.status_code == 200
    emp_headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    fields_resp = client.get("/fields", headers=emp_headers)
    assert fields_resp.status_code == 200
    field_id = fields_resp.json()[0]["id"]

    submit_resp = client.post(
        "/submissions",
        json={
            "field_id": field_id,
            "date": "2026-07-31",
            "time": "08:30:00",
            "meters": FOUR_METER_PAYLOAD,
        },
        headers=emp_headers,
    )
    assert submit_resp.status_code == 422



def test_export_endpoints_require_admin(client):
    resp = client.get("/admin/submissions/export/excel")
    assert resp.status_code == 401



def test_export_excel_and_pdf(client, admin_token):
    headers = {"Authorization": f"Bearer {admin_token}"}
    excel_resp = client.get("/admin/submissions/export/excel", headers=headers)
    assert excel_resp.status_code == 200
    assert excel_resp.headers["content-type"].startswith("application/vnd.openxmlformats")

    workbook = load_workbook(io.BytesIO(excel_resp.content))
    sheet = workbook.active
    assert sheet.title == "Daily TL Counts"
    assert sheet["A1"].value == "Field"
    assert sheet["M1"].value == "Overall Avg"
    assert sheet["A2"].value == FIELD_PAYLOAD["name"]
    if sheet.max_row > 1:
        assert sheet["E2"].value == 1
        assert sheet["L3"].value == 16
        assert sheet["E4"].value == "AVG (n=2)"
        assert sheet["F4"].value == 6.5
        assert sheet["L4"].value == 12.5
        assert sheet["M4"].value == 66.5

    pdf_resp = client.get("/admin/submissions/export/pdf", headers=headers)
    assert pdf_resp.status_code == 200
    assert pdf_resp.headers["content-type"] == "application/pdf"



def test_admin_can_reset_employee_password(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    users_resp = client.get("/admin/users", headers=admin_headers)
    assert users_resp.status_code == 200
    employee = next(u for u in users_resp.json() if u["username"] == "ion_test")

    reset_resp = client.patch(
        f"/admin/users/{employee['id']}/password",
        json={"password": "brandnewpass"},
        headers=admin_headers,
    )
    assert reset_resp.status_code == 200

    old_login = client.post("/auth/login", json={"username": "ion_test", "password": "parola123"})
    assert old_login.status_code == 401

    new_login = client.post("/auth/login", json={"username": "ion_test", "password": "brandnewpass"})
    assert new_login.status_code == 200


def test_reset_password_requires_admin(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}
    users_resp = client.get("/admin/users", headers=admin_headers)
    employee = next(u for u in users_resp.json() if u["username"] == "ion_test")

    login_resp = client.post("/auth/login", json={"username": "ion_test", "password": "brandnewpass"})
    emp_headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    resp = client.patch(
        f"/admin/users/{employee['id']}/password",
        json={"password": "shouldnotwork"},
        headers=emp_headers,
    )
    assert resp.status_code == 403


def test_reset_password_unknown_user(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}
    resp = client.patch(
        "/admin/users/999999/password",
        json={"password": "whatever123"},
        headers=admin_headers,
    )
    assert resp.status_code == 404


def test_admin_can_delete_employee_and_submissions(client, admin_token):
    admin_headers = {"Authorization": f"Bearer {admin_token}"}

    create_resp = client.post(
        "/admin/users",
        json={"username": "delete_me", "password": "parola123", "full_name": "Delete Me", "role": "employee"},
        headers=admin_headers,
    )
    assert create_resp.status_code == 201
    employee_id = create_resp.json()["id"]

    helper_resp = client.post(
        "/admin/users",
        json={"username": "helper_emp", "password": "helper123", "role": "employee"},
        headers=admin_headers,
    )
    assert helper_resp.status_code == 201

    login_resp = client.post("/auth/login", json={"username": "delete_me", "password": "parola123"})
    assert login_resp.status_code == 200
    delete_headers = {"Authorization": f"Bearer {login_resp.json()['access_token']}"}

    fields_resp = client.get("/fields", headers=delete_headers)
    assert fields_resp.status_code == 200
    field_id = fields_resp.json()[0]["id"]

    submit_resp = client.post(
        "/submissions",
        json={
            "field_id": field_id,
            "date": "2026-08-01",
            "meters": [METER_PAYLOAD[0]],
        },
        headers=delete_headers,
    )
    assert submit_resp.status_code == 201
    submission_id = submit_resp.json()["id"]

    helper_login = client.post("/auth/login", json={"username": "helper_emp", "password": "helper123"})
    assert helper_login.status_code == 200
    helper_headers = {"Authorization": f"Bearer {helper_login.json()['access_token']}"}

    forbidden_resp = client.delete(f"/admin/users/{employee_id}", headers=helper_headers)
    assert forbidden_resp.status_code == 403

    missing_resp = client.delete("/admin/users/999999", headers=admin_headers)
    assert missing_resp.status_code == 404

    users_resp = client.get("/admin/users", headers=admin_headers)
    admin_user = next(user for user in users_resp.json() if user["username"] == "admin")
    admin_delete_resp = client.delete(f"/admin/users/{admin_user['id']}", headers=admin_headers)
    assert admin_delete_resp.status_code == 400
    assert admin_delete_resp.json()["detail"] == "Only employee accounts can be deleted"

    delete_resp = client.delete(f"/admin/users/{employee_id}", headers=admin_headers)
    assert delete_resp.status_code == 204

    users_after_delete = client.get("/admin/users", headers=admin_headers)
    assert all(user["id"] != employee_id for user in users_after_delete.json())

    submissions_after_delete = client.get("/admin/submissions", headers=admin_headers)
    assert all(submission["id"] != submission_id for submission in submissions_after_delete.json())
